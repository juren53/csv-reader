#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
CSV Reader Application

A PyQt6-based application to view CSV and XLSX files with data displayed vertically.
Features:
- Display CSV/XLSX records vertically with header as labels
- Navigate records using left/right arrow keys
- Recent files menu
- Zoom functionality with Ctrl + mouse wheel
- Support for Excel files (.xlsx) using openpyxl

Changelog:
v0.0.5 2026-01-09 10:45 CST - Added dynamic header row selection feature
"""

import sys
import os
import csv
try:
    import openpyxl
    XLSX_SUPPORT_AVAILABLE = True
except ImportError:
    XLSX_SUPPORT_AVAILABLE = False
    openpyxl = None
from PyQt6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, 
                           QHBoxLayout, QGridLayout, QLabel, QScrollArea, 
                           QFileDialog, QMessageBox, QMenu, QSizePolicy,
                           QTableWidget, QTableWidgetItem, QHeaderView, QAbstractItemView,
                           QLineEdit, QPushButton)
from PyQt6.QtCore import Qt, QSettings, QSize, QEvent
from PyQt6.QtGui import QFont, QKeySequence, QFontMetrics, QPalette, QAction, QColor, QIcon

class TableView(QTableWidget):
    """Widget to display CSV data in a spreadsheet-like table format"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        
        # Configure table properties
        self.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)  # Read-only
        self.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectItems)
        self.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)
        self.setWordWrap(True)
        
        # Enable scrolling
        self.setHorizontalScrollMode(QAbstractItemView.ScrollMode.ScrollPerPixel)
        self.setVerticalScrollMode(QAbstractItemView.ScrollMode.ScrollPerPixel)
        
        # Configure headers
        self.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self.horizontalHeader().setStretchLastSection(True)
        self.horizontalHeader().setDefaultAlignment(Qt.AlignmentFlag.AlignLeft)
        self.verticalHeader().setDefaultSectionSize(25)
        self.verticalHeader().setVisible(True)
        
        # Default font settings
        self.base_font_size = 10
        self.current_zoom_level = 100  # 100%
        self.updateFontSize()
        
        # Enable sorting
        self.setSortingEnabled(True)
        
        # Search tracking
        self.search_results = []  # List of (row, col) tuples
        self.current_search_index = -1
        self.search_text = ""
        
    def updateFontSize(self):
        """Update the font size based on the current zoom level"""
        font_size = int(self.base_font_size * (self.current_zoom_level / 100))
        font = QFont()
        font.setPointSize(max(6, font_size))  # Ensure minimum readable size
        self.setFont(font)
        
        # Update row heights to match font
        self.verticalHeader().setDefaultSectionSize(max(25, font_size + 10))
        
    def zoom(self, delta):
        """Change zoom level by a percentage"""
        # Adjust zoom level with limits (40% to 300%)
        new_zoom = max(40, min(300, self.current_zoom_level + delta))
        if new_zoom != self.current_zoom_level:
            self.current_zoom_level = new_zoom
            print(f"Zoom level changed to {self.current_zoom_level}%")  # Debug output
            self.updateFontSize()
            return True
        return False
        
    def displayData(self, headers, data):
        """Display CSV data in the table"""
        if not headers or not data:
            self.clear()
            return
            
        # Set table dimensions
        self.setRowCount(len(data))
        self.setColumnCount(len(headers))
        
        # Set headers
        self.setHorizontalHeaderLabels(headers)
        
        # Populate table with data
        for row_idx, row_data in enumerate(data):
            # Ensure row has enough columns
            if len(row_data) < len(headers):
                row_data = row_data + [''] * (len(headers) - len(row_data))
                
            for col_idx, cell_value in enumerate(row_data):
                item = QTableWidgetItem(str(cell_value))
                item.setTextAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignTop)
                self.setItem(row_idx, col_idx, item)
        
        # Resize columns to content
        self.resizeColumnsToContents()
        
        # Update font size
        self.updateFontSize()
        
        # Clear search highlights
        self.clearSearch()
    
    def searchTable(self, search_text, case_sensitive=False):
        """Search for text in the table and highlight results"""
        # Clear previous search
        self.clearSearch()
        
        if not search_text:
            return 0
        
        self.search_text = search_text
        self.search_results = []
        
        # Search through all cells
        for row in range(self.rowCount()):
            for col in range(self.columnCount()):
                item = self.item(row, col)
                if item:
                    cell_text = item.text()
                    # Check if search text is in cell
                    if case_sensitive:
                        if search_text in cell_text:
                            self.search_results.append((row, col))
                    else:
                        if search_text.lower() in cell_text.lower():
                            self.search_results.append((row, col))
        
        # Highlight all results
        for row, col in self.search_results:
            item = self.item(row, col)
            if item:
                item.setBackground(QColor(255, 255, 0, 100))  # Light yellow
        
        # Move to first result if any
        if self.search_results:
            self.current_search_index = 0
            self.highlightCurrentResult()
        
        return len(self.search_results)
    
    def highlightCurrentResult(self):
        """Highlight the current search result with a different color"""
        if not self.search_results or self.current_search_index < 0:
            return
        
        # Unhighlight previous result (back to light yellow)
        for i, (row, col) in enumerate(self.search_results):
            item = self.item(row, col)
            if item:
                if i == self.current_search_index:
                    item.setBackground(QColor(255, 165, 0, 150))  # Orange for current
                else:
                    item.setBackground(QColor(255, 255, 0, 100))  # Light yellow for others
        
        # Scroll to current result
        if self.search_results:
            row, col = self.search_results[self.current_search_index]
            self.scrollToItem(self.item(row, col))
    
    def nextSearchResult(self):
        """Move to next search result"""
        if not self.search_results:
            return False
        
        self.current_search_index = (self.current_search_index + 1) % len(self.search_results)
        self.highlightCurrentResult()
        return True
    
    def previousSearchResult(self):
        """Move to previous search result"""
        if not self.search_results:
            return False
        
        self.current_search_index = (self.current_search_index - 1) % len(self.search_results)
        self.highlightCurrentResult()
        return True
    
    def clearSearch(self):
        """Clear search highlights"""
        for row, col in self.search_results:
            item = self.item(row, col)
            if item:
                item.setBackground(QColor(255, 255, 255, 0))  # Transparent
        
        self.search_results = []
        self.current_search_index = -1
        self.search_text = ""

class RecordView(QScrollArea):
    """Widget to display a single CSV record vertically"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.container = QWidget()
        self.grid_layout = QGridLayout(self.container)
        self.setWidget(self.container)
        self.setWidgetResizable(True)
        
        # Set up the widget appearance
        self.setFrameShape(QScrollArea.Shape.NoFrame)
        self.container.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        
        # Default font settings
        self.base_font_size = 10
        self.current_zoom_level = 100  # 100%
        self.updateFontSize()
        
        # Styling
        self.label_style = "font-weight: bold; padding-right: 10px;"

    def updateFontSize(self):
        """Update the font size based on the current zoom level"""
        font_size = int(self.base_font_size * (self.current_zoom_level / 100))
        font = QFont()
        font.setPointSize(max(6, font_size))  # Ensure minimum readable size
        
        # Apply font to all labels
        for i in range(self.grid_layout.count()):
            item = self.grid_layout.itemAt(i).widget()
            if isinstance(item, QLabel):
                item.setFont(font)
                
        # Update the container to reflect the changes
        self.container.adjustSize()
        self.container.update()  # Force a visual update

    def zoom(self, delta):
        """Change zoom level by a percentage"""
        # Adjust zoom level with limits (40% to 300%)
        new_zoom = max(40, min(300, self.current_zoom_level + delta))
        if new_zoom != self.current_zoom_level:
            self.current_zoom_level = new_zoom
            print(f"Zoom level changed to {self.current_zoom_level}%")  # Debug output
            self.updateFontSize()
            # Force the scroll area to refresh
            self.setWidgetResizable(False)
            self.setWidgetResizable(True)
            return True
        return False
        
    def displayRecord(self, headers, record):
        """Display a record with headers as labels"""
        # Clear previous content
        while self.grid_layout.count():
            item = self.grid_layout.takeAt(0)
            widget = item.widget()
            if widget:
                widget.deleteLater()
                
        if not headers or not record:
            return
            
        # Display each field with its header
        for row, (header, value) in enumerate(zip(headers, record)):
            # Create label for header
            header_label = QLabel(header)
            header_label.setStyleSheet(self.label_style)
            header_label.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignTop)
            
            # Create label for value
            value_label = QLabel(value)
            value_label.setTextInteractionFlags(Qt.TextInteractionFlag.TextSelectableByMouse)
            value_label.setWordWrap(True)
            value_label.setAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignTop)
            
            # Add to layout
            self.grid_layout.addWidget(header_label, row, 0)
            self.grid_layout.addWidget(value_label, row, 1)
            
        # Set column stretch
        self.grid_layout.setColumnStretch(0, 0)  # Header column
        self.grid_layout.setColumnStretch(1, 1)  # Value column
        
        # Update font size
        self.updateFontSize()

class CSVReaderApp(QMainWindow):
    """Main application window for CSV Reader"""
    
    MAX_RECENT_FILES = 10
    VERSION = "v0.0.5  2026-01-09  10:45 CST"
    
    def __init__(self):
        super().__init__()
        
        # Application settings
        self.settings = QSettings("CSVReader", "csv-reader")
        
        # Initialize data structures
        self.csv_data = []
        self.headers = []
        self.current_record_index = 0
        self.current_file_path = None
        self.current_view_mode = "record"  # "record" or "table"
        self.search_visible = False
        
        # Set up the UI
        self.initUI()
        
        # Load recent files list
        self.updateRecentFilesMenu()
        
        # Install event filter for key events at application level
        QApplication.instance().installEventFilter(self)
    
    def initUI(self):
        """Initialize the user interface"""
        # Set window properties
        self.setWindowTitle("CSV/XLSX Reader")
        self.setMinimumSize(800, 600)
        
        # Set application icon
        icon_path = os.path.join(os.path.dirname(__file__), "icons", "ICON_csv-reader.png")
        if os.path.exists(icon_path):
            self.setWindowIcon(QIcon(icon_path))
        
        # Create central widget and layout
        central_widget = QWidget()
        main_layout = QVBoxLayout(central_widget)
        
        # Create view toggle controls
        view_controls_layout = QHBoxLayout()
        
        # View mode buttons
        self.record_view_btn = QPushButton("Record View")
        self.record_view_btn.setCheckable(True)
        self.record_view_btn.setChecked(True)
        self.record_view_btn.clicked.connect(self.switchToRecordView)
        
        self.table_view_btn = QPushButton("Table View")
        self.table_view_btn.setCheckable(True)
        self.table_view_btn.setChecked(False)
        self.table_view_btn.clicked.connect(self.switchToTableView)
        
        view_controls_layout.addWidget(self.record_view_btn)
        view_controls_layout.addWidget(self.table_view_btn)
        view_controls_layout.addStretch()
        
        # Search button (only visible in table view)
        self.search_toggle_btn = QPushButton("Search")
        self.search_toggle_btn.setCheckable(True)
        self.search_toggle_btn.setChecked(False)
        self.search_toggle_btn.clicked.connect(self.toggleSearch)
        self.search_toggle_btn.hide()  # Hidden initially
        view_controls_layout.addWidget(self.search_toggle_btn)
        
        main_layout.addLayout(view_controls_layout)
        
        # Create search bar (initially hidden)
        self.search_widget = QWidget()
        search_layout = QHBoxLayout(self.search_widget)
        search_layout.setContentsMargins(0, 0, 0, 0)
        
        search_layout.addWidget(QLabel("Search:"))
        
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Enter search text...")
        self.search_input.returnPressed.connect(self.performSearch)
        search_layout.addWidget(self.search_input)
        
        self.search_btn = QPushButton("Find")
        self.search_btn.clicked.connect(self.performSearch)
        search_layout.addWidget(self.search_btn)
        
        self.prev_result_btn = QPushButton("Previous")
        self.prev_result_btn.clicked.connect(self.previousSearchResult)
        search_layout.addWidget(self.prev_result_btn)
        
        self.next_result_btn = QPushButton("Next")
        self.next_result_btn.clicked.connect(self.nextSearchResult)
        search_layout.addWidget(self.next_result_btn)
        
        self.search_result_label = QLabel("")
        search_layout.addWidget(self.search_result_label)
        
        self.clear_search_btn = QPushButton("Clear")
        self.clear_search_btn.clicked.connect(self.clearTableSearch)
        search_layout.addWidget(self.clear_search_btn)
        
        self.search_widget.hide()
        main_layout.addWidget(self.search_widget)
        
        # Create views container
        self.views_container = QWidget()
        self.views_layout = QVBoxLayout(self.views_container)
        
        # Create both views
        self.record_view = RecordView()
        self.table_view = TableView()
        
        # Connect table row click to record view
        self.table_view.cellClicked.connect(self.onTableCellClicked)
        
        # Initially show only record view
        self.views_layout.addWidget(self.record_view)
        self.views_layout.addWidget(self.table_view)
        self.table_view.hide()
        
        main_layout.addWidget(self.views_container)
        
        # Add navigation controls (only for record view)
        self.nav_widget = QWidget()
        nav_layout = QHBoxLayout(self.nav_widget)
        
        prev_btn_widget = QPushButton("< Previous")
        prev_btn_widget.clicked.connect(self.previousRecord)
        
        next_btn_widget = QPushButton("Next >")
        next_btn_widget.clicked.connect(self.nextRecord)
        
        nav_layout.addWidget(prev_btn_widget)
        
        # Add navigation info label
        self.nav_label = QLabel("No file loaded")
        self.nav_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        nav_layout.addWidget(self.nav_label)
        
        nav_layout.addWidget(next_btn_widget)
        main_layout.addWidget(self.nav_widget)
        
        self.setCentralWidget(central_widget)
        
        # Create menu bar
        menubar = self.menuBar()
        
        # File menu
        file_menu = menubar.addMenu('&File')
        
        # Open action
        open_action = QAction('&Open...', self)
        open_action.setShortcut(QKeySequence.StandardKey.Open)
        open_action.triggered.connect(self.openFile)
        file_menu.addAction(open_action)
        
        # Recent files submenu
        self.recent_files_menu = QMenu('&Recent Files', self)
        file_menu.addMenu(self.recent_files_menu)
        
        # Separator
        file_menu.addSeparator()
        
        # Exit action
        exit_action = QAction('E&xit', self)
        exit_action.setShortcut(QKeySequence.StandardKey.Quit)
        exit_action.triggered.connect(self.close)
        file_menu.addAction(exit_action)
        
        # View menu
        view_menu = menubar.addMenu('&View')
        
        # Record view action
        record_view_action = QAction('&Record View', self)
        record_view_action.setCheckable(True)
        record_view_action.setChecked(True)
        record_view_action.triggered.connect(self.switchToRecordView)
        view_menu.addAction(record_view_action)
        self.record_view_action = record_view_action
        
        # Table view action
        table_view_action = QAction('&Table View', self)
        table_view_action.setCheckable(True)
        table_view_action.setChecked(False)
        table_view_action.triggered.connect(self.switchToTableView)
        view_menu.addAction(table_view_action)
        self.table_view_action = table_view_action
        
        # Separator
        view_menu.addSeparator()
        
        # Toggle view action
        toggle_view_action = QAction('&Toggle View', self)
        toggle_view_action.setShortcut(QKeySequence('Ctrl+T'))
        toggle_view_action.triggered.connect(self.toggleView)
        view_menu.addAction(toggle_view_action)
        
        # Help menu
        help_menu = menubar.addMenu('&Help')
        
        # Quick Reference action
        quick_ref_action = QAction('&Quick Reference', self)
        quick_ref_action.setShortcut(QKeySequence('F1'))
        quick_ref_action.triggered.connect(self.showQuickReference)
        help_menu.addAction(quick_ref_action)
        
        # Changelog action
        changelog_action = QAction('&Changelog', self)
        changelog_action.triggered.connect(self.showChangelog)
        help_menu.addAction(changelog_action)
        
        # Separator
        help_menu.addSeparator()
        
        # About action
        about_action = QAction('&About CSV Reader', self)
        about_action.triggered.connect(self.showAbout)
        help_menu.addAction(about_action)
        
        # Status bar
        self.statusBar().showMessage(f'Ready | {self.VERSION}')
        
        # Center window on screen
        screen_size = QApplication.primaryScreen().size()
        self.move((screen_size.width() - self.width()) // 2,
                  (screen_size.height() - self.height()) // 2)
    
    def switchToRecordView(self):
        """Switch to record view mode"""
        if self.current_view_mode == "record":
            return
            
        self.current_view_mode = "record"
        self.record_view.show()
        self.table_view.hide()
        self.nav_widget.show()
        
        # Hide search controls
        self.search_toggle_btn.hide()
        self.search_widget.hide()
        self.search_visible = False
        
        # Update UI state
        self.record_view_btn.setChecked(True)
        self.table_view_btn.setChecked(False)
        self.record_view_action.setChecked(True)
        self.table_view_action.setChecked(False)
        
        # Update status bar
        self.updateStatusBar()
        
        # Display current record if data is loaded
        if self.csv_data and self.headers:
            self.displayCurrentRecord()
    
    def switchToTableView(self):
        """Switch to table view mode"""
        if self.current_view_mode == "table":
            return
            
        self.current_view_mode = "table"
        self.record_view.hide()
        self.table_view.show()
        self.nav_widget.hide()
        
        # Show search button
        self.search_toggle_btn.show()
        
        # Update UI state
        self.record_view_btn.setChecked(False)
        self.table_view_btn.setChecked(True)
        self.record_view_action.setChecked(False)
        self.table_view_action.setChecked(True)
        
        # Update status bar
        self.updateStatusBar()
        
        # Display table data if data is loaded
        if self.csv_data and self.headers:
            self.table_view.displayData(self.headers, self.csv_data)
    
    def toggleView(self):
        """Toggle between record and table view"""
        if self.current_view_mode == "record":
            self.switchToTableView()
        else:
            self.switchToRecordView()
    
    def onTableCellClicked(self, row, column):
        """Handle table cell click - switch to record view for that row"""
        if not self.csv_data:
            return
            
        # Update current record index to the clicked row
        self.current_record_index = row
        
        # Switch to record view
        self.switchToRecordView()
    
    def updateStatusBar(self):
        """Update status bar with current view mode and statistics"""
        if not self.csv_data or not self.headers:
            self.statusBar().showMessage(f'Ready | {self.VERSION}')
            return
            
        if self.current_view_mode == "record":
            self.statusBar().showMessage(f"Record View - Record {self.current_record_index + 1} of {len(self.csv_data)} | {self.VERSION}")
        else:
            self.statusBar().showMessage(f"Table View - {len(self.csv_data)} rows, {len(self.headers)} columns | {self.VERSION}")
    
    def updateRecentFilesMenu(self):
        """Update the recent files menu with stored recent files"""
        self.recent_files_menu.clear()
        
        recent_files = self.settings.value("recentFiles", [])
        if not recent_files:
            recent_files = []
        
        for i, file_path in enumerate(recent_files):
            if i < self.MAX_RECENT_FILES and os.path.exists(file_path):
                action = QAction(f"&{i+1} {file_path}", self)
                action.setData(file_path)
                action.triggered.connect(lambda checked, file_path=file_path: self.openRecentFile(file_path))
                self.recent_files_menu.addAction(action)
                
        # Add clear action if there are recent files
        if self.recent_files_menu.actions():
            self.recent_files_menu.addSeparator()
            clear_action = QAction('&Clear Recent Files', self)
            clear_action.triggered.connect(self.clearRecentFiles)
            self.recent_files_menu.addAction(clear_action)
    
    def addToRecentFiles(self, file_path):
        """Add a file to the recent files list"""
        if not file_path:
            return
            
        recent_files = self.settings.value("recentFiles", [])
        if not recent_files:
            recent_files = []
            
        # Remove if already in the list
        if file_path in recent_files:
            recent_files.remove(file_path)
            
        # Add to the front of the list
        recent_files.insert(0, file_path)
        
        # Limit list size
        if len(recent_files) > self.MAX_RECENT_FILES:
            recent_files = recent_files[:self.MAX_RECENT_FILES]
            
        # Save updated list
        self.settings.setValue("recentFiles", recent_files)
        
        # Update the menu
        self.updateRecentFilesMenu()
    
    def clearRecentFiles(self):
        """Clear the recent files list"""
        self.settings.setValue("recentFiles", [])
        self.updateRecentFilesMenu()
    
    def loadLastViewedFile(self):
        """Load the last viewed data file on startup"""
        last_file = self.settings.value("lastViewedFile", "")
        if last_file and os.path.exists(last_file):
            self.loadDataFile(last_file)
    
    def toggleSearch(self):
        """Toggle search bar visibility"""
        self.search_visible = not self.search_visible
        if self.search_visible:
            self.search_widget.show()
            self.search_input.setFocus()
        else:
            self.search_widget.hide()
            self.clearTableSearch()
    
    def performSearch(self):
        """Perform search in table view"""
        search_text = self.search_input.text()
        if not search_text:
            return
        
        count = self.table_view.searchTable(search_text, case_sensitive=False)
        if count > 0:
            self.search_result_label.setText(f"Found {count} match(es) - Showing 1/{count}")
        else:
            self.search_result_label.setText("No matches found")
    
    def nextSearchResult(self):
        """Move to next search result"""
        if self.table_view.nextSearchResult():
            total = len(self.table_view.search_results)
            current = self.table_view.current_search_index + 1
            self.search_result_label.setText(f"Found {total} match(es) - Showing {current}/{total}")
    
    def previousSearchResult(self):
        """Move to previous search result"""
        if self.table_view.previousSearchResult():
            total = len(self.table_view.search_results)
            current = self.table_view.current_search_index + 1
            self.search_result_label.setText(f"Found {total} match(es) - Showing {current}/{total}")
    
    def clearTableSearch(self):
        """Clear search in table view"""
        self.search_input.clear()
        self.search_result_label.setText("")
        self.table_view.clearSearch()
    
    def showQuickReference(self):
        """Display quick reference dialog"""
        quick_ref_text = """
<h2>CSV/XLSX Reader - Quick Reference</h2>

<h3>Supported File Formats</h3>
<ul>
<li><b>CSV Files</b>: Standard comma-separated values files (.csv)</li>
<li><b>Excel Files</b>: Excel workbook files (.xlsx) - first sheet only</li>
</ul>

<h3>View Modes</h3>
<ul>
<li><b>Record View</b>: Displays one data record at a time with fields shown vertically (header-value pairs)</li>
<li><b>Table View</b>: Displays all data in a spreadsheet-like table format</li>
</ul>

<h3>Navigation (Record View)</h3>
<ul>
<li><b>Left Arrow</b>: Previous record</li>
<li><b>Right Arrow</b>: Next record</li>
<li><b>H</b>: Select current record as header row</li>
<li><b>Click Previous/Next buttons</b>: Navigate between records</li>
</ul>

<h3>Table View</h3>
<ul>
<li><b>Click any cell</b>: Switch to Record View for that row</li>
<li><b>Click column headers</b>: Sort by that column</li>
</ul>

<h3>Keyboard Shortcuts</h3>
<ul>
<li><b>Ctrl+O</b>: Open CSV file</li>
<li><b>Ctrl+T</b>: Toggle between Record and Table views</li>
<li><b>Ctrl+MouseWheel</b>: Zoom in/out (40%-300%)</li>
<li><b>F1</b>: Show Quick Reference</li>
<li><b>Ctrl+Q</b>: Quit application</li>
</ul>

<h3>Features</h3>
<ul>
<li><b>Recent Files</b>: Access recently opened CSV files from File menu</li>
<li><b>Auto-load</b>: Last viewed file opens automatically on startup</li>
<li><b>Independent Zoom</b>: Each view mode maintains its own zoom level</li>
<li><b>Text Selection</b>: Select and copy text in Record View</li>
<li><b>Dynamic Header Selection</b>: Choose any row as the header in Record View</li>
</ul>
        """
        
        msg = QMessageBox(self)
        msg.setWindowTitle("Quick Reference")
        msg.setTextFormat(Qt.TextFormat.RichText)
        msg.setText(quick_ref_text)
        msg.setIcon(QMessageBox.Icon.Information)
        msg.setStandardButtons(QMessageBox.StandardButton.Ok)
        msg.exec()
    
    def showChangelog(self):
        """Display changelog dialog"""
        changelog_path = os.path.join(os.path.dirname(__file__), "CHANGELOG.md")
        changelog_text = ""
        
        try:
            if os.path.exists(changelog_path):
                with open(changelog_path, 'r', encoding='utf-8') as f:
                    changelog_text = f.read()
            else:
                changelog_text = "CHANGELOG.md file not found."
        except Exception as e:
            changelog_text = f"Error reading CHANGELOG.md: {str(e)}"
        
        # Convert markdown to basic HTML for display
        changelog_html = changelog_text.replace('\n', '<br>').replace('**', '<b>').replace('**', '</b>')
        
        msg = QMessageBox(self)
        msg.setWindowTitle("Changelog")
        msg.setTextFormat(Qt.TextFormat.RichText)
        msg.setText(changelog_html)
        msg.setIcon(QMessageBox.Icon.Information)
        msg.setStandardButtons(QMessageBox.StandardButton.Ok)
        msg.exec()
    
    def showAbout(self):
        """Display about dialog"""
        about_text = f"""
<h2>CSV/XLSX Reader</h2>
<p><b>Version:</b> {self.VERSION}</p>
<p>A PyQt6-based application for viewing CSV and XLSX files with both record and table view modes.</p>
<p><b>Features:</b></p>
<ul>
<li>Support for CSV and XLSX files</li>
<li>Multi-sheet XLSX files (loads first sheet)</li>
<li>Dual view modes (Record and Table)</li>
<li>Dynamic header row selection</li>
<li>Keyboard navigation and shortcuts</li>
<li>Zoom functionality (40%-300%)</li>
<li>Recent files tracking</li>
<li>Auto-load last viewed file</li>
</ul>
<p><b>Copyright © 2026</b></p>
        """
        
        msg = QMessageBox(self)
        msg.setWindowTitle("About CSV/XLSX Reader")
        msg.setTextFormat(Qt.TextFormat.RichText)
        msg.setText(about_text)
        msg.setIcon(QMessageBox.Icon.Information)
        msg.setStandardButtons(QMessageBox.StandardButton.Ok)
        msg.exec()
    
    def openRecentFile(self, file_path):
        """Open a file from the recent files menu"""
        if os.path.exists(file_path):
            self.loadDataFile(file_path)
        else:
            QMessageBox.warning(self, "File Not Found", 
                               f"The file {file_path} does not exist.",
                               QMessageBox.StandardButton.Ok)
            
            # Remove from recent files
            recent_files = self.settings.value("recentFiles", [])
            if file_path in recent_files:
                recent_files.remove(file_path)
                self.settings.setValue("recentFiles", recent_files)
                self.updateRecentFilesMenu()
    
    def openFile(self):
        """Open a file dialog to select a CSV or XLSX file"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Open Data File", "",
            "CSV Files (*.csv);;Excel Files (*.xlsx);;All Files (*)"
        )

        if file_path:
            self.loadDataFile(file_path)

    def _loadCSVData(self, file_path):
        """Parse CSV file and populate headers and csv_data"""
        with open(file_path, 'r', newline='', encoding='utf-8') as f:
            reader = csv.reader(f)
            self.csv_data = list(reader)

        if not self.csv_data:
            raise Exception("File is empty")

        # Extract headers (first row)
        self.headers = self.csv_data[0]

        # Remove headers from data
        self.csv_data = self.csv_data[1:]

        if not self.csv_data:
            raise Exception("No data records found")

    def _loadXLSXData(self, file_path):
        """Parse XLSX file and populate headers and csv_data"""
        if not XLSX_SUPPORT_AVAILABLE:
            raise Exception(
                "XLSX support requires the 'openpyxl' library.\n\n"
                "Install it using: pip install openpyxl"
            )

        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        except Exception as e:
            raise Exception(f"Failed to open XLSX file: {str(e)}")

        if len(workbook.sheetnames) == 0:
            raise Exception("XLSX file contains no sheets")

        sheet = workbook[workbook.sheetnames[0]]

        # Convert sheet to list of lists
        data = []
        for row in sheet.iter_rows(values_only=True):
            row_data = [str(cell) if cell is not None else '' for cell in row]
            data.append(row_data)

        workbook.close()

        if not data:
            raise Exception("First sheet is empty")

        # Extract headers (first row)
        self.headers = data[0]

        # Remove headers from data
        self.csv_data = data[1:]

        if not self.csv_data:
            raise Exception("No data records found (only headers present)")

    def loadDataFile(self, file_path):
        """Load and parse a CSV or XLSX file"""
        try:
            # Detect file type
            file_ext = os.path.splitext(file_path)[1].lower()

            if file_ext == '.xlsx':
                self._loadXLSXData(file_path)
            elif file_ext == '.csv':
                self._loadCSVData(file_path)
            else:
                raise Exception(f"Unsupported file type: {file_ext}")

            # Reset current record index
            self.current_record_index = 0

            # Update UI
            self.current_file_path = file_path
            self.setWindowTitle(f"CSV/XLSX Reader - {file_path}")

            # Display data based on current view mode
            if self.current_view_mode == "record":
                self.displayCurrentRecord()
            else:
                self.table_view.displayData(self.headers, self.csv_data)

            # Add to recent files
            self.addToRecentFiles(file_path)

            # Save as last viewed file
            self.settings.setValue("lastViewedFile", file_path)

            # Update status
            self.updateStatusBar()

        except Exception as e:
            QMessageBox.critical(self, "Error",
                               f"Failed to load file: {str(e)}",
                               QMessageBox.StandardButton.Ok)
            self.statusBar().showMessage("Failed to load file")

    def loadCSVFile(self, file_path):
        """Load and parse a CSV file (deprecated - use loadDataFile)"""
        self.loadDataFile(file_path)
    
    def displayCurrentRecord(self):
        """Display the current record or table based on view mode"""
        if not self.csv_data or not self.headers:
            self.nav_label.setText("No file loaded")
            return
            
        if self.current_view_mode == "record":
            # Display the current record
            record = self.csv_data[self.current_record_index]
            
            # Extend record if shorter than headers
            if len(record) < len(self.headers):
                record = record + [''] * (len(self.headers) - len(record))
                
            # Update the record view
            self.record_view.displayRecord(self.headers, record)
            
            # Update navigation label
            self.nav_label.setText(f"Record {self.current_record_index + 1} of {len(self.csv_data)}")
        else:
            # Update table view
            self.table_view.displayData(self.headers, self.csv_data)
        
        # Update status bar
        self.updateStatusBar()
    
    def previousRecord(self):
        """Go to the previous record"""
        if not self.csv_data:
            return
            
        if self.current_record_index > 0:
            self.current_record_index -= 1
            self.displayCurrentRecord()
    
    def nextRecord(self):
        """Go to the next record"""
        if not self.csv_data:
            return

        if self.current_record_index < len(self.csv_data) - 1:
            self.current_record_index += 1
            self.displayCurrentRecord()

    def selectCurrentRecordAsHeader(self):
        """Allow user to select current record as the new header row"""
        if not self.csv_data or not self.headers:
            return

        # Don't allow if only one row (would leave no data)
        if len(self.csv_data) <= 1:
            QMessageBox.warning(
                self,
                "Cannot Change Header",
                "Cannot set header when only one data row exists.",
                QMessageBox.StandardButton.Ok
            )
            return

        current_record = self.csv_data[self.current_record_index]

        # Show confirmation dialog
        if self.showHeaderSelectionDialog(current_record, self.current_record_index):
            self.reassignHeaderRow(self.current_record_index)

    def showHeaderSelectionDialog(self, record_data, record_index):
        """Show dialog to confirm header selection with horizontal display"""
        from PyQt6.QtWidgets import QDialog, QVBoxLayout, QTableWidget, \
                                     QTableWidgetItem, QDialogButtonBox, QLabel

        dialog = QDialog(self)
        dialog.setWindowTitle("Select Header Row")
        dialog.setMinimumWidth(600)

        layout = QVBoxLayout()

        # Add explanation label
        info_label = QLabel(
            f"Select record {record_index + 1} as the new header row?\n\n"
            "The current header will become a data row.\n"
            "All rows before this will remain as data."
        )
        info_label.setWordWrap(True)
        layout.addWidget(info_label)

        # Create table to show current headers and selected row
        preview_table = QTableWidget(2, len(self.headers))
        preview_table.setHorizontalHeaderLabels([f"Col {i+1}" for i in range(len(self.headers))])
        preview_table.setVerticalHeaderLabels(["Current Header", "New Header"])

        # Populate table
        for col_idx, header in enumerate(self.headers):
            preview_table.setItem(0, col_idx, QTableWidgetItem(str(header)))

        for col_idx, value in enumerate(record_data):
            if col_idx < len(self.headers):
                preview_table.setItem(1, col_idx, QTableWidgetItem(str(value)))

        preview_table.resizeColumnsToContents()
        preview_table.setMaximumHeight(150)
        layout.addWidget(preview_table)

        # Add buttons
        button_box = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Yes | QDialogButtonBox.StandardButton.No
        )
        button_box.accepted.connect(dialog.accept)
        button_box.rejected.connect(dialog.reject)
        layout.addWidget(button_box)

        dialog.setLayout(layout)

        # Show dialog and return result
        return dialog.exec() == QDialog.DialogCode.Accepted

    def reassignHeaderRow(self, new_header_index):
        """Reassign the header row to a different row in the data"""
        if new_header_index < 0 or new_header_index >= len(self.csv_data):
            return

        # Get the new header row
        new_header_row = self.csv_data[new_header_index]

        # Pad new header if needed to match maximum row length
        max_cols = max(len(row) for row in self.csv_data)
        if len(new_header_row) < max_cols:
            new_header_row = list(new_header_row) + \
                            [f"Column{i+1}" for i in range(len(new_header_row), max_cols)]
        else:
            new_header_row = list(new_header_row)

        # Build new data structure
        new_data = []

        # 1. Add old header as first data row
        new_data.append(self.headers)

        # 2. Add all rows before the selected row
        new_data.extend(self.csv_data[0:new_header_index])

        # 3. Add all rows after the selected row (skip the selected row itself)
        new_data.extend(self.csv_data[new_header_index + 1:])

        # Normalize all rows to have same length as headers
        for i in range(len(new_data)):
            if len(new_data[i]) < len(new_header_row):
                new_data[i] = list(new_data[i]) + [''] * (len(new_header_row) - len(new_data[i]))

        # Update application state
        self.headers = new_header_row
        self.csv_data = new_data

        # Adjust current record index if needed
        if self.current_record_index >= len(self.csv_data):
            self.current_record_index = len(self.csv_data) - 1
        if self.current_record_index < 0:
            self.current_record_index = 0

        # Update both views
        if self.current_view_mode == "record":
            self.displayCurrentRecord()
        else:
            self.table_view.displayData(self.headers, self.csv_data)
            self.switchToRecordView()  # Switch back to record view after change

        # Update UI
        self.updateStatusBar()
        self.statusBar().showMessage("Header row updated", 3000)

    def keyPressEvent(self, event):
        """Handle key press events for navigation"""
        if not self.csv_data:
            return super().keyPressEvent(event)
            
        if event.key() == Qt.Key.Key_Left:
            # Go to previous record
            self.previousRecord()
            event.accept()
                
        elif event.key() == Qt.Key.Key_Right:
            # Go to next record
            self.nextRecord()
            event.accept()
                
        else:
            super().keyPressEvent(event)
            
    def eventFilter(self, obj, event):
        """Global event filter to capture key events regardless of focus"""
        if event.type() == QEvent.Type.KeyPress:
            # Handle left/right arrow keys for navigation in record view only
            if self.current_view_mode == "record":
                if event.key() == Qt.Key.Key_Left:
                    self.previousRecord()
                    return True
                elif event.key() == Qt.Key.Key_Right:
                    self.nextRecord()
                    return True
                elif event.key() == Qt.Key.Key_H:
                    self.selectCurrentRecordAsHeader()
                    return True
        elif event.type() == QEvent.Type.Wheel:
            # Handle wheel events with Ctrl for zooming
            if event.modifiers() & Qt.KeyboardModifier.ControlModifier:
                delta = 15 if event.angleDelta().y() > 0 else -15
                
                # Zoom the appropriate view
                if self.current_view_mode == "record":
                    if self.record_view.zoom(delta):
                        self.statusBar().showMessage(f"Zoom: {self.record_view.current_zoom_level}%", 2000)
                else:
                    if self.table_view.zoom(delta):
                        self.statusBar().showMessage(f"Zoom: {self.table_view.current_zoom_level}%", 2000)
                return True
                
        # Standard event processing
        return super().eventFilter(obj, event)
    
    def wheelEvent(self, event):
        """Handle mouse wheel events for zooming with Ctrl modifier"""
        if event.modifiers() & Qt.KeyboardModifier.ControlModifier:
            # Calculate zoom delta based on wheel direction (15% steps)
            delta = 15 if event.angleDelta().y() > 0 else -15
            
            # Apply zoom to appropriate view
            if self.current_view_mode == "record":
                if self.record_view.zoom(delta):
                    self.statusBar().showMessage(f"Zoom: {self.record_view.current_zoom_level}%", 2000)
            else:
                if self.table_view.zoom(delta):
                    self.statusBar().showMessage(f"Zoom: {self.table_view.current_zoom_level}%", 2000)
                
            # Accept the event
            event.accept()
            return  # Prevent further event processing
        else:
            # Default handling for normal scrolling
            super().wheelEvent(event)

def main():
    """Main application entry point"""
    app = QApplication(sys.argv)
    
    # Set application name for settings
    app.setApplicationName("CSVReader")
    app.setOrganizationName("CSVReader")
    
    # Create and show the main window
    window = CSVReaderApp()
    window.show()
    
    # Check for command line arguments (file to open)
    if len(sys.argv) > 1:
        file_path = sys.argv[1]
        if os.path.exists(file_path):
            file_ext = file_path.lower()
            if file_ext.endswith('.csv') or file_ext.endswith('.xlsx'):
                window.loadDataFile(file_path)
            else:
                print(f"Error: Unsupported file type. Please provide a .csv or .xlsx file.")
    else:
        # Only load last viewed file if no command line argument
        window.loadLastViewedFile()
    
    # Start the event loop
    sys.exit(app.exec())

if __name__ == "__main__":
    main()

